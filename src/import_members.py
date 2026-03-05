"""
import_members.py — Load KAFAMemberList into DynamoDB kopera-member table

Usage:
    python import_members.py                              # uses KAFAMemberList_complete.xlsx in current dir
    python import_members.py --file path/to/file.xlsx    # custom file path
    python import_members.py --dry-run                   # preview without writing to DynamoDB

Requirements:
    pip install boto3 openpyxl
"""

import argparse
import sys
import boto3
from decimal import Decimal
from openpyxl import load_workbook

# ── Config ────────────────────────────────────────────────────────────────────

TABLE_NAME  = "kopera-member"
AWS_REGION  = "us-east-1"

# Column order must match the xlsx exactly
COLUMNS = [
    "memberId",       # A
    "full_name",      # B
    "date_of_birth",  # C
    "id_number",      # D
    "id_type",        # E
    "address",        # F
    "nationality",    # G
    "phone",          # H
    "issued_date",    # I  (DOM — Date of Membership)
    "companyId",      # J
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def clean(value) -> str:
    """Coerce any cell value to a clean string. Skip None/empty."""
    if value is None:
        return ""
    return str(value).strip()


def build_item(row: tuple) -> dict | None:
    """
    Map a spreadsheet row to a DynamoDB item.
    Returns None if the row is missing the required keys (memberId, companyId).
    """
    item = {}
    for idx, col_name in enumerate(COLUMNS):
        raw = row[idx] if idx < len(row) else None
        val = clean(raw)
        if val:
            item[col_name] = val

    # Primary key must be present
    if not item.get("memberId") or not item.get("companyId"):
        return None

    return item


def load_xlsx(filepath: str) -> list[dict]:
    """Read all data rows from the xlsx and return list of DynamoDB items."""
    wb   = load_workbook(filepath, data_only=True)
    ws   = wb.active
    items = []
    skipped = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        item = build_item(row)
        if item is None:
            skipped.append(row_idx)
            continue

        items.append(item)

    if skipped:
        print(f"⚠️  Skipped {len(skipped)} rows with missing memberId or companyId: rows {skipped}")

    return items


def batch_write(table, items: list[dict], dry_run: bool) -> tuple[int, int]:
    """
    Write items to DynamoDB using batch_writer (handles retries automatically).
    Returns (success_count, error_count).
    """
    success = 0
    errors  = 0

    if dry_run:
        for item in items:
            print(f"  [DRY RUN] Would write: {item}")
        return len(items), 0

    with table.batch_writer() as batch:
        for item in items:
            try:
                batch.put_item(Item=item)
                success += 1
            except Exception as exc:
                print(f"  ❌ Failed to write {item.get('memberId')}: {exc}")
                errors += 1

    return success, errors


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import KAFA members from xlsx into DynamoDB")
    parser.add_argument("--file",    default="KAFAMemberList_complete.xlsx", help="Path to xlsx file")
    parser.add_argument("--table",   default=TABLE_NAME,                     help="DynamoDB table name")
    parser.add_argument("--region",  default=AWS_REGION,                     help="AWS region")
    parser.add_argument("--dry-run", action="store_true",                    help="Preview without writing")
    args = parser.parse_args()

    print(f"\n{'='*55}")
    print(f"  KAFA Member Import")
    print(f"  File   : {args.file}")
    print(f"  Table  : {args.table}")
    print(f"  Region : {args.region}")
    print(f"  Mode   : {'DRY RUN — no data will be written' if args.dry_run else 'LIVE'}")
    print(f"{'='*55}\n")

    # ── Load xlsx ─────────────────────────────────────────────────────────────
    print(f"📂 Reading {args.file} ...")
    try:
        items = load_xlsx(args.file)
    except FileNotFoundError:
        print(f"❌ File not found: {args.file}")
        sys.exit(1)

    print(f"✅ {len(items)} member records loaded from xlsx\n")

    if not items:
        print("Nothing to import. Exiting.")
        sys.exit(0)

    # Preview first 3 records
    print("Preview (first 3 records):")
    for item in items[:3]:
        print(f"  {item}")
    print()

    # ── Connect to DynamoDB ───────────────────────────────────────────────────
    dynamodb = boto3.resource("dynamodb", region_name=args.region)
    table    = dynamodb.Table(args.table)

    if not args.dry_run:
        # Confirm before writing
        confirm = input(f"⚠️  About to write {len(items)} records to '{args.table}'. Proceed? [y/N]: ")
        if confirm.strip().lower() != "y":
            print("Aborted.")
            sys.exit(0)

    # ── Write to DynamoDB ─────────────────────────────────────────────────────
    print(f"\n{'Writing' if not args.dry_run else 'Simulating'} {len(items)} records ...")
    success, errors = batch_write(table, items, dry_run=args.dry_run)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'='*55}")
    print(f"  Import complete")
    print(f"  ✅ Written  : {success}")
    print(f"  ❌ Errors   : {errors}")
    print(f"{'='*55}\n")

    if errors > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
